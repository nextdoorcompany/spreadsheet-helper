import openpyxl

def read_with_headers(filename, sheetname, header_row=1, start_col=1):
	header_row = int(header_row)
	start_col = int(start_col)
	result = []
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_sheet_by_name(sheetname)
	c = start_col
	headers = {}
	while sheet.cell(row=header_row, column=c).value:
		header_val = str(sheet.cell(row=header_row, column=c).value)
		headers[c] = header_val.replace('\n', '').replace('\\', '').replace("'", '').replace('"', '').replace(' ', '')
		c += 1
	max_col = c
	max_row = sheet.max_row
	for row in range(header_row + 1, max_row + 1):
		row_data = {}
		for col in range(start_col, max_col):
			val = sheet.cell(row=row, column=col).value
			if val is None:
				val = ''
			row_data[headers[col]] = val
		result.append(row_data)	
	return result